# -*- coding: utf-8 -*-
"""
Portal Interactive Features Controller
Provides JSON endpoints for AJAX-powered portal widgets
"""

from odoo import http, fields, _
from odoo.http import request
from odoo.addons.portal.controllers.portal import CustomerPortal
import json


class PortalInteractiveController(CustomerPortal):
    """
    Controller for interactive portal features:
    - AJAX pagination and filtering
    - Real-time price calculation
    - Barcode scanning
    - Export functionality
    """

    # NOTE: calculate_retrieval_price moved to portal.py to avoid duplicate routes
    
    def _get_customer_rates(self, partner):
        """
        Get pricing rates for a customer.
        
        Lookup order:
        1. Customer negotiated rate (active)
        2. Base rate (company default)
        3. Hardcoded fallbacks
        """
        rates = {}
        
        # Try customer negotiated rate first
        negotiated = request.env['customer.negotiated.rate'].sudo().search([
            ('partner_id', '=', partner.id),
            ('state', '=', 'active'),
        ], limit=1)
        
        if negotiated:
            rates = {
                'document_retrieval_rate': negotiated.per_service_rate or 0,
                'delivery_rate': negotiated.per_service_rate or 0,  # Could add specific field
                'indexing_rate': negotiated.per_document_rate or 0,
                'per_container_fee': 5.00,  # Could add to negotiated rates model
            }
            # Return if we have a retrieval rate
            if rates.get('document_retrieval_rate'):
                return rates
        
        # Fall back to base rates
        base_rate = request.env['base.rate'].sudo().search([
            ('company_id', '=', request.env.company.id),
            ('state', '=', 'active'),
        ], limit=1)
        
        if base_rate:
            rates = {
                'document_retrieval_rate': base_rate.document_retrieval_rate or 25.00,
                'delivery_rate': base_rate.delivery_rate or 50.00,
                'indexing_rate': base_rate.indexing_rate or 1.50,
                'per_container_fee': 5.00,  # Could add to base rates model
                'pickup_rate': base_rate.pickup_rate or 0,
                'destruction_rate': base_rate.destruction_rate or 0,
                'scanning_rate': base_rate.scanning_rate or 0,
            }
            return rates
        
        # Ultimate fallback - hardcoded defaults
        return {
            'document_retrieval_rate': 25.00,
            'delivery_rate': 50.00,
            'indexing_rate': 1.50,
            'per_container_fee': 5.00,
            'pickup_rate': 125.00,
            'destruction_rate': 12.50,
            'scanning_rate': 0.65,
        }

    # NOTE: process_barcode route moved to portal.py to avoid duplicate
    # The route /my/barcode/process/<string:operation> is in portal.py

    @http.route(['/my/containers/export'], type='http', auth="user", website=True)
    def export_containers(self, format='xlsx', container_ids=None, **kw):
        """
        Export container list to Excel/CSV
        
        :param format: Export format (xlsx, csv)
        :param container_ids: Comma-separated list of container IDs to export (optional)
        :return: File download
        """
        try:
            Container = request.env['records.container']
            
            # If specific container IDs provided, use those
            if container_ids:
                ids = [int(x) for x in container_ids.split(',') if x.strip().isdigit()]
                # Security: Only allow exporting user's own containers
                domain = [
                    ('id', 'in', ids),
                    ('partner_id', '=', request.env.user.partner_id.id)
                ]
                containers = Container.search(domain, order='name')
            else:
                # Otherwise use standard domain filter
                domain = self._get_containers_domain(**kw)
                containers = Container.search(domain, order='name')
            
            if format == 'xlsx':
                return self._export_containers_xlsx(containers)
            elif format == 'csv':
                return self._export_containers_csv(containers)
            else:
                return request.redirect('/my/containers')
                
        except Exception as e:
            return request.redirect('/my/containers?error=' + str(e))

    def _get_containers_domain(self, search=None, filterby=None, **kw):
        """Build domain for container search"""
        domain = [('partner_id', '=', request.env.user.partner_id.id)]
        
        if search:
            domain += ['|', '|',
                      ('name', 'ilike', search),
                      ('description', 'ilike', search),
                      ('location_id.name', 'ilike', search)]
        
        if filterby and filterby != 'all':
            if filterby == 'active':
                domain.append(('state', '=', 'active'))
            elif filterby == 'destroyed':
                domain.append(('state', '=', 'destroyed'))
            elif filterby == 'pending':
                domain.append(('state', '=', 'pending'))
        
        return domain

    def _export_containers_xlsx(self, containers):
        """Export containers to Excel with comprehensive data"""
        import io
        from odoo.tools.misc import xlsxwriter
        from datetime import datetime
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Containers')
        
        # Header format
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4F81BD',
            'font_color': 'white',
            'border': 1,
            'text_wrap': True,
            'valign': 'vcenter'
        })
        
        # Date format
        date_format = workbook.add_format({'num_format': 'yyyy-mm-dd'})
        
        # Write headers - comprehensive columns
        headers = [
            'Container Number',
            'Barcode', 
            'Status',
            'Location',
            'Department',
            'Description',
            'Storage Start Date',
            'Destruction Due Date',
            'Retention Policy',
            'File Count',
            'Document Count',
            'Date Created',
            'Last Updated'
        ]
        
        # Set column widths
        col_widths = [18, 15, 12, 20, 20, 30, 15, 18, 20, 10, 12, 15, 15]
        for col, (header, width) in enumerate(zip(headers, col_widths)):
            worksheet.write(0, col, header, header_format)
            worksheet.set_column(col, col, width)
        
        # Write data
        for row, container in enumerate(containers, start=1):
            worksheet.write(row, 0, container.name or '')
            worksheet.write(row, 1, container.barcode or container.temp_barcode or '')
            worksheet.write(row, 2, dict(container._fields['state'].selection).get(container.state, container.state) if container.state else '')
            worksheet.write(row, 3, container.location_id.name if container.location_id else '')
            worksheet.write(row, 4, container.department_id.name if container.department_id else '')
            worksheet.write(row, 5, container.description or '')
            worksheet.write(row, 6, fields.Date.to_string(container.storage_start_date) if container.storage_start_date else '')
            worksheet.write(row, 7, fields.Date.to_string(container.destruction_due_date) if container.destruction_due_date else '')
            worksheet.write(row, 8, container.retention_policy_id.name if container.retention_policy_id else '')
            worksheet.write(row, 9, container.file_count if hasattr(container, 'file_count') else 0)
            worksheet.write(row, 10, container.document_count if hasattr(container, 'document_count') else 0)
            worksheet.write(row, 11, fields.Date.to_string(container.create_date.date()) if container.create_date else '')
            worksheet.write(row, 12, fields.Date.to_string(container.write_date.date()) if container.write_date else '')
        
        # Freeze header row
        worksheet.freeze_panes(1, 0)
        
        workbook.close()
        output.seek(0)
        
        # Generate filename with date
        filename = f"containers_export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return request.make_response(
            output.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
        )

    def _export_containers_csv(self, containers):
        """Export containers to CSV with comprehensive data"""
        import csv
        import io
        from datetime import datetime
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write comprehensive header
        writer.writerow([
            'Container Number',
            'Barcode', 
            'Status',
            'Location',
            'Department',
            'Description',
            'Storage Start Date',
            'Destruction Due Date',
            'Retention Policy',
            'File Count',
            'Document Count',
            'Date Created',
            'Last Updated'
        ])
        
        # Write data
        for container in containers:
            writer.writerow([
                container.name or '',
                container.barcode or container.temp_barcode or '',
                dict(container._fields['state'].selection).get(container.state, container.state) if container.state else '',
                container.location_id.name if container.location_id else '',
                container.department_id.name if container.department_id else '',
                container.description or '',
                fields.Date.to_string(container.storage_start_date) if container.storage_start_date else '',
                fields.Date.to_string(container.destruction_due_date) if container.destruction_due_date else '',
                container.retention_policy_id.name if container.retention_policy_id else '',
                container.file_count if hasattr(container, 'file_count') else 0,
                container.document_count if hasattr(container, 'document_count') else 0,
                fields.Date.to_string(container.create_date.date()) if container.create_date else '',
                fields.Date.to_string(container.write_date.date()) if container.write_date else ''
            ])
        
        # Generate filename with date
        filename = f"containers_export_{datetime.now().strftime('%Y-%m-%d')}.csv"
        
        return request.make_response(
            output.getvalue(),
            headers=[
                ('Content-Type', 'text/csv'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
        )

    @http.route(['/my/requests/export'], type='http', auth="user", website=True)
    def export_requests(self, format='xlsx', **kw):
        """Export requests list to Excel/CSV"""
        try:
            Request = request.env['portal.request']
            requests = Request.search([
                ('partner_id', '=', request.env.user.partner_id.id)
            ], order='create_date desc')
            
            if format == 'xlsx':
                return self._export_requests_xlsx(requests)
            else:
                return self._export_requests_csv(requests)
                
        except Exception as e:
            return request.redirect('/my/requests?error=' + str(e))

    def _export_requests_xlsx(self, requests):
        """Export requests to Excel"""
        import io
        from odoo.tools.misc import xlsxwriter
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Requests')
        
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4F81BD',
            'font_color': 'white',
            'border': 1
        })
        
        headers = ['Request #', 'Type', 'Status', 'Description', 'Created', 'Updated']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        
        for row, req in enumerate(requests, start=1):
            worksheet.write(row, 0, req.name or '')
            worksheet.write(row, 1, req.request_type or '')
            worksheet.write(row, 2, req.state or '')
            worksheet.write(row, 3, req.description or '')
            worksheet.write(row, 4, fields.Date.to_string(req.create_date.date()) if req.create_date else '')
            worksheet.write(row, 5, fields.Date.to_string(req.write_date.date()) if req.write_date else '')
        
        workbook.close()
        output.seek(0)
        
        return request.make_response(
            output.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', 'attachment; filename="requests_export.xlsx"')
            ]
        )

    def _export_requests_csv(self, requests):
        """Export requests to CSV"""
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['Request #', 'Type', 'Status', 'Description', 'Created', 'Updated'])
        
        for req in requests:
            writer.writerow([
                req.name or '',
                req.request_type or '',
                req.state or '',
                req.description or '',
                fields.Date.to_string(req.create_date.date()) if req.create_date else '',
                fields.Date.to_string(req.write_date.date()) if req.write_date else ''
            ])
        
        return request.make_response(
            output.getvalue(),
            headers=[
                ('Content-Type', 'text/csv'),
                ('Content-Disposition', 'attachment; filename="requests_export.csv"')
            ]
        )

    # ================================================================
    # FILE FOLDER EXPORT
    # ================================================================

    @http.route(['/my/files/export'], type='http', auth="user", website=True)
    def export_files(self, format='xlsx', file_ids=None, **kw):
        """Export file folders list to Excel/CSV with comprehensive columns"""
        try:
            File = request.env['records.file']
            
            # If specific file IDs provided, use those
            if file_ids:
                ids = [int(x) for x in file_ids.split(',') if x.strip().isdigit()]
                # Security: Only allow exporting user's own files
                domain = [
                    ('id', 'in', ids),
                    ('partner_id', '=', request.env.user.partner_id.id)
                ]
                files = File.search(domain, order='name')
            else:
                # Otherwise use standard domain filter  
                domain = [('partner_id', '=', request.env.user.partner_id.id)]
                if kw.get('search'):
                    domain += ['|', '|',
                              ('name', 'ilike', kw['search']),
                              ('description', 'ilike', kw['search']),
                              ('barcode', 'ilike', kw['search'])]
                files = File.search(domain, order='name')
            
            if format == 'xlsx':
                return self._export_files_xlsx(files)
            elif format == 'csv':
                return self._export_files_csv(files)
            else:
                return request.redirect('/my/inventory/files')
                
        except Exception as e:
            return request.redirect('/my/inventory/files?error=' + str(e))

    def _export_files_xlsx(self, files):
        """Export file folders to Excel with comprehensive data"""
        import io
        from odoo.tools.misc import xlsxwriter
        from datetime import datetime
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('File Folders')
        
        # Header format
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4F81BD',
            'font_color': 'white',
            'border': 1,
            'text_wrap': True,
            'valign': 'vcenter'
        })
        
        # Write headers - comprehensive columns
        headers = [
            'File Name/Number',
            'Barcode', 
            'Container',
            'Location',
            'Department',
            'Description',
            'Document Count',
            'Responsible Person',
            'Date Created',
            'Last Updated'
        ]
        
        # Set column widths
        col_widths = [25, 20, 20, 20, 20, 35, 12, 20, 15, 15]
        for col, (header, width) in enumerate(zip(headers, col_widths)):
            worksheet.write(0, col, header, header_format)
            worksheet.set_column(col, col, width)
        
        # Write data
        for row, file_rec in enumerate(files, start=1):
            worksheet.write(row, 0, file_rec.name or '')
            worksheet.write(row, 1, file_rec.barcode or file_rec.temp_barcode or '')
            worksheet.write(row, 2, file_rec.container_id.name if file_rec.container_id else '')
            worksheet.write(row, 3, file_rec.current_location_id.name if file_rec.current_location_id else (file_rec.container_id.current_location_id.name if file_rec.container_id and file_rec.container_id.current_location_id else ''))
            worksheet.write(row, 4, file_rec.department_id.name if file_rec.department_id else '')
            worksheet.write(row, 5, file_rec.description or '')
            worksheet.write(row, 6, file_rec.document_count or 0)
            worksheet.write(row, 7, file_rec.responsible_person_id.name if file_rec.responsible_person_id else '')
            worksheet.write(row, 8, fields.Date.to_string(file_rec.create_date.date()) if file_rec.create_date else '')
            worksheet.write(row, 9, fields.Date.to_string(file_rec.write_date.date()) if file_rec.write_date else '')
        
        # Freeze header row
        worksheet.freeze_panes(1, 0)
        
        workbook.close()
        output.seek(0)
        
        # Generate filename with date
        filename = f"file_folders_export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return request.make_response(
            output.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
        )

    def _export_files_csv(self, files):
        """Export file folders to CSV with comprehensive data"""
        import csv
        import io
        from datetime import datetime
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write comprehensive header
        writer.writerow([
            'File Name/Number',
            'Barcode', 
            'Container',
            'Location',
            'Department',
            'Description',
            'Document Count',
            'Responsible Person',
            'Date Created',
            'Last Updated'
        ])
        
        # Write data
        for file_rec in files:
            writer.writerow([
                file_rec.name or '',
                file_rec.barcode or file_rec.temp_barcode or '',
                file_rec.container_id.name if file_rec.container_id else '',
                file_rec.current_location_id.name if file_rec.current_location_id else (file_rec.container_id.current_location_id.name if file_rec.container_id and file_rec.container_id.current_location_id else ''),
                file_rec.department_id.name if file_rec.department_id else '',
                file_rec.description or '',
                file_rec.document_count or 0,
                file_rec.responsible_person_id.name if file_rec.responsible_person_id else '',
                fields.Date.to_string(file_rec.create_date.date()) if file_rec.create_date else '',
                fields.Date.to_string(file_rec.write_date.date()) if file_rec.write_date else ''
            ])
        
        # Generate filename with date
        filename = f"file_folders_export_{datetime.now().strftime('%Y-%m-%d')}.csv"
        
        return request.make_response(
            output.getvalue(),
            headers=[
                ('Content-Type', 'text/csv'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
        )

    # ================================================================
    # USER EXPORT (Company Portal Admin)
    # ================================================================

    @http.route(['/my/users/export'], type='http', auth="user", website=True)
    def export_users(self, format='xlsx', **kw):
        """Export users list to Excel/CSV (company admin only)"""
        # Security: Only company admins can export
        if not request.env.user.has_group('records_management.group_portal_company_admin'):
            return request.redirect('/my/home')
        
        try:
            partner = request.env.user.partner_id.commercial_partner_id
            
            # Get all portal users for this company
            users = request.env['res.users'].sudo().search([
                ('partner_id.parent_id', '=', partner.id),
                ('groups_id', 'in', [request.env.ref('base.group_portal').id])
            ], order='name')
            
            # Get departments for user department mapping
            departments = request.env['records.department'].sudo().search([
                ('company_id', '=', partner.id)
            ])
            dept_map = {d.id: d.name for d in departments}
            
            if format == 'xlsx':
                return self._export_users_xlsx(users, dept_map)
            elif format == 'csv':
                return self._export_users_csv(users, dept_map)
            else:
                return request.redirect('/my/users')
                
        except Exception as e:
            return request.redirect('/my/users?error=' + str(e))

    def _export_users_xlsx(self, users, dept_map):
        """Export users to Excel with comprehensive data"""
        import io
        from odoo.tools.misc import xlsxwriter
        from datetime import datetime
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Portal Users')
        
        # Header format
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4F81BD',
            'font_color': 'white',
            'border': 1,
            'text_wrap': True,
            'valign': 'vcenter'
        })
        
        # Write headers
        headers = [
            'Name',
            'Email (Login)', 
            'Department',
            'Phone',
            'Job Title',
            'Status',
            'Last Login',
            'Date Created'
        ]
        
        # Set column widths
        col_widths = [25, 30, 20, 15, 20, 10, 18, 15]
        for col, (header, width) in enumerate(zip(headers, col_widths)):
            worksheet.write(0, col, header, header_format)
            worksheet.set_column(col, col, width)
        
        # Write data
        for row, user in enumerate(users, start=1):
            dept_name = ''
            if hasattr(user.partner_id, 'department_id') and user.partner_id.department_id:
                dept_name = user.partner_id.department_id.name
            
            worksheet.write(row, 0, user.name or '')
            worksheet.write(row, 1, user.login or '')
            worksheet.write(row, 2, dept_name)
            worksheet.write(row, 3, user.partner_id.phone or user.partner_id.mobile or '')
            worksheet.write(row, 4, user.partner_id.function or '')
            worksheet.write(row, 5, 'Active' if user.active else 'Inactive')
            worksheet.write(row, 6, fields.Datetime.to_string(user.login_date) if user.login_date else 'Never')
            worksheet.write(row, 7, fields.Date.to_string(user.create_date.date()) if user.create_date else '')
        
        # Freeze header row
        worksheet.freeze_panes(1, 0)
        
        workbook.close()
        output.seek(0)
        
        # Generate filename with date
        filename = f"portal_users_export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return request.make_response(
            output.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
        )

    def _export_users_csv(self, users, dept_map):
        """Export users to CSV with comprehensive data"""
        import csv
        import io
        from datetime import datetime
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Name',
            'Email (Login)', 
            'Department',
            'Phone',
            'Job Title',
            'Status',
            'Last Login',
            'Date Created'
        ])
        
        # Write data
        for user in users:
            dept_name = ''
            if hasattr(user.partner_id, 'department_id') and user.partner_id.department_id:
                dept_name = user.partner_id.department_id.name
            
            writer.writerow([
                user.name or '',
                user.login or '',
                dept_name,
                user.partner_id.phone or user.partner_id.mobile or '',
                user.partner_id.function or '',
                'Active' if user.active else 'Inactive',
                fields.Datetime.to_string(user.login_date) if user.login_date else 'Never',
                fields.Date.to_string(user.create_date.date()) if user.create_date else ''
            ])
        
        # Generate filename with date
        filename = f"portal_users_export_{datetime.now().strftime('%Y-%m-%d')}.csv"
        
        return request.make_response(
            output.getvalue(),
            headers=[
                ('Content-Type', 'text/csv'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
        )

    # ================================================================
    # BULK IMPORT FUNCTIONALITY
    # ================================================================

    @http.route(['/my/import/template/<string:data_type>'], type='http', auth="user", website=True)
    def download_import_template(self, data_type, format='xlsx', **kw):
        """Download a blank import template for containers, files, or users"""
        try:
            if data_type == 'containers':
                return self._generate_container_template(format)
            elif data_type == 'files':
                return self._generate_file_template(format)
            elif data_type == 'users':
                # Security: Only company admins can download user template
                if not request.env.user.has_group('records_management.group_portal_company_admin'):
                    return request.redirect('/my/home')
                return self._generate_user_template(format)
            else:
                return request.redirect('/my/home')
        except Exception as e:
            return request.redirect('/my/home?error=' + str(e))

    def _generate_container_template(self, format='xlsx'):
        """Generate blank container import template"""
        import io
        from odoo.tools.misc import xlsxwriter
        from datetime import datetime
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Containers')
        
        # Header format
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4F81BD',
            'font_color': 'white',
            'border': 1
        })
        
        # Instructions format
        instruction_format = workbook.add_format({
            'italic': True,
            'font_color': '#666666',
            'text_wrap': True
        })
        
        # Headers with instructions
        headers = [
            ('Container Number*', 'Required. Unique identifier for the container (e.g., BOX-2024-001)'),
            ('Description', 'Optional. Brief description of contents'),
            ('Department', 'Optional. Department name - must match existing department'),
            ('Retention Policy', 'Optional. Retention policy name - must match existing policy'),
            ('Destruction Due Date', 'Optional. Format: YYYY-MM-DD'),
            ('Notes', 'Optional. Additional notes')
        ]
        
        for col, (header, instruction) in enumerate(headers):
            worksheet.write(0, col, header, header_format)
            worksheet.write(1, col, instruction, instruction_format)
            worksheet.set_column(col, col, 25)
        
        # Add sample data row
        worksheet.write(2, 0, 'BOX-2024-001')
        worksheet.write(2, 1, 'HR Personnel Files 2024')
        worksheet.write(2, 2, 'Human Resources')
        worksheet.write(2, 3, '7 Years')
        worksheet.write(2, 4, '2031-12-31')
        worksheet.write(2, 5, 'Confidential')
        
        workbook.close()
        output.seek(0)
        
        return request.make_response(
            output.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', 'attachment; filename="container_import_template.xlsx"')
            ]
        )

    def _generate_file_template(self, format='xlsx'):
        """Generate blank file folder import template"""
        import io
        from odoo.tools.misc import xlsxwriter
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('File Folders')
        
        # Header format
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4F81BD',
            'font_color': 'white',
            'border': 1
        })
        
        # Instructions format
        instruction_format = workbook.add_format({
            'italic': True,
            'font_color': '#666666',
            'text_wrap': True
        })
        
        # Headers with instructions
        headers = [
            ('File Name/Number*', 'Required. Name or number for the file (e.g., John Smith or CASE-2024-001)'),
            ('Container Number', 'Optional. Container this file belongs to - must match existing container'),
            ('Description', 'Optional. Brief description of file contents'),
            ('Department', 'Optional. Department name - must match existing department'),
            ('Notes', 'Optional. Additional notes')
        ]
        
        for col, (header, instruction) in enumerate(headers):
            worksheet.write(0, col, header, header_format)
            worksheet.write(1, col, instruction, instruction_format)
            worksheet.set_column(col, col, 30)
        
        # Add sample data row
        worksheet.write(2, 0, 'John Smith Personnel File')
        worksheet.write(2, 1, 'BOX-2024-001')
        worksheet.write(2, 2, 'Employee personnel records')
        worksheet.write(2, 3, 'Human Resources')
        worksheet.write(2, 4, '')
        
        workbook.close()
        output.seek(0)
        
        return request.make_response(
            output.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', 'attachment; filename="file_folder_import_template.xlsx"')
            ]
        )

    def _generate_user_template(self, format='xlsx'):
        """Generate blank user import template"""
        import io
        from odoo.tools.misc import xlsxwriter
        
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Portal Users')
        
        # Header format
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4F81BD',
            'font_color': 'white',
            'border': 1
        })
        
        # Instructions format
        instruction_format = workbook.add_format({
            'italic': True,
            'font_color': '#666666',
            'text_wrap': True
        })
        
        # Headers with instructions
        headers = [
            ('Name*', 'Required. Full name of the user'),
            ('Email*', 'Required. Email address (used as login)'),
            ('Department*', 'Required. Department name - must match existing department'),
            ('Phone', 'Optional. Phone number'),
            ('Job Title', 'Optional. User\'s job title or position')
        ]
        
        for col, (header, instruction) in enumerate(headers):
            worksheet.write(0, col, header, header_format)
            worksheet.write(1, col, instruction, instruction_format)
            worksheet.set_column(col, col, 30)
        
        # Add sample data row
        worksheet.write(2, 0, 'Jane Doe')
        worksheet.write(2, 1, 'jane.doe@example.com')
        worksheet.write(2, 2, 'Human Resources')
        worksheet.write(2, 3, '555-123-4567')
        worksheet.write(2, 4, 'HR Manager')
        
        workbook.close()
        output.seek(0)
        
        return request.make_response(
            output.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', 'attachment; filename="user_import_template.xlsx"')
            ]
        )

    @http.route(['/my/import/containers'], type='http', auth="user", website=True, methods=['POST'], csrf=True)
    def import_containers(self, **post):
        """Import containers from uploaded file"""
        import base64
        import io
        
        try:
            file_data = post.get('file')
            if not file_data:
                return request.redirect('/my/inventory/containers?error=no_file')
            
            # Read file content
            file_content = file_data.read()
            
            # Determine file type and parse
            if file_data.filename.endswith('.xlsx'):
                result = self._parse_container_xlsx(file_content)
            elif file_data.filename.endswith('.csv'):
                result = self._parse_container_csv(file_content)
            else:
                return request.redirect('/my/inventory/containers?error=invalid_format')
            
            if result.get('error'):
                return request.redirect('/my/inventory/containers?error=' + result['error'])
            
            return request.redirect(f'/my/inventory/containers?imported={result.get("count", 0)}')
            
        except Exception as e:
            return request.redirect('/my/inventory/containers?error=' + str(e))

    def _parse_container_xlsx(self, file_content):
        """Parse container data from XLSX file"""
        import io
        try:
            import openpyxl
        except ImportError:
            return {'error': 'Excel support not available'}
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            worksheet = workbook.active
            
            partner = request.env.user.partner_id
            Container = request.env['records.container']
            Department = request.env['records.department']
            RetentionPolicy = request.env['records.retention.policy']
            
            created_count = 0
            
            # Skip header rows (row 1 = headers, row 2 = instructions)
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=3, values_only=True), start=3):
                name = row[0]
                if not name:  # Skip empty rows
                    continue
                
                description = row[1] if len(row) > 1 else ''
                department_name = row[2] if len(row) > 2 else ''
                retention_name = row[3] if len(row) > 3 else ''
                destruction_date = row[4] if len(row) > 4 else None
                notes = row[5] if len(row) > 5 else ''
                
                # Find department by name
                department_id = False
                if department_name:
                    dept = Department.sudo().search([
                        ('name', 'ilike', department_name),
                        ('company_id', '=', partner.commercial_partner_id.id)
                    ], limit=1)
                    if dept:
                        department_id = dept.id
                
                # Find retention policy by name
                retention_id = False
                if retention_name:
                    policy = RetentionPolicy.sudo().search([
                        ('name', 'ilike', retention_name),
                        ('active', '=', True)
                    ], limit=1)
                    if policy:
                        retention_id = policy.id
                
                # Create container
                vals = {
                    'name': str(name),
                    'description': str(description) if description else '',
                    'partner_id': partner.id,
                    'department_id': department_id,
                    'retention_policy_id': retention_id,
                    'created_via_portal': True,
                    'storage_start_date': fields.Date.today(),
                }
                
                if destruction_date:
                    try:
                        if isinstance(destruction_date, str):
                            vals['destruction_due_date'] = destruction_date
                        else:
                            vals['destruction_due_date'] = fields.Date.to_string(destruction_date)
                    except Exception:
                        pass
                
                Container.sudo().create(vals)
                created_count += 1
            
            return {'count': created_count}
            
        except Exception as e:
            return {'error': str(e)}

    def _parse_container_csv(self, file_content):
        """Parse container data from CSV file"""
        import csv
        import io
        
        try:
            content = file_content.decode('utf-8')
            reader = csv.reader(io.StringIO(content))
            
            partner = request.env.user.partner_id
            Container = request.env['records.container']
            Department = request.env['records.department']
            RetentionPolicy = request.env['records.retention.policy']
            
            created_count = 0
            
            # Skip header rows
            next(reader, None)  # Header row
            next(reader, None)  # Instructions row
            
            for row in reader:
                if not row or not row[0]:  # Skip empty rows
                    continue
                
                name = row[0]
                description = row[1] if len(row) > 1 else ''
                department_name = row[2] if len(row) > 2 else ''
                retention_name = row[3] if len(row) > 3 else ''
                destruction_date = row[4] if len(row) > 4 else ''
                
                # Find department by name
                department_id = False
                if department_name:
                    dept = Department.sudo().search([
                        ('name', 'ilike', department_name),
                        ('company_id', '=', partner.commercial_partner_id.id)
                    ], limit=1)
                    if dept:
                        department_id = dept.id
                
                # Find retention policy by name
                retention_id = False
                if retention_name:
                    policy = RetentionPolicy.sudo().search([
                        ('name', 'ilike', retention_name),
                        ('active', '=', True)
                    ], limit=1)
                    if policy:
                        retention_id = policy.id
                
                # Create container
                vals = {
                    'name': name,
                    'description': description,
                    'partner_id': partner.id,
                    'department_id': department_id,
                    'retention_policy_id': retention_id,
                    'created_via_portal': True,
                    'storage_start_date': fields.Date.today(),
                }
                
                if destruction_date:
                    vals['destruction_due_date'] = destruction_date
                
                Container.sudo().create(vals)
                created_count += 1
            
            return {'count': created_count}
            
        except Exception as e:
            return {'error': str(e)}

    @http.route(['/my/import/files'], type='http', auth="user", website=True, methods=['POST'], csrf=True)
    def import_files(self, **post):
        """Import file folders from uploaded file"""
        try:
            file_data = post.get('file')
            if not file_data:
                return request.redirect('/my/inventory/files?error=no_file')
            
            # Read file content
            file_content = file_data.read()
            
            # Determine file type and parse
            if file_data.filename.endswith('.xlsx'):
                result = self._parse_file_xlsx(file_content)
            elif file_data.filename.endswith('.csv'):
                result = self._parse_file_csv(file_content)
            else:
                return request.redirect('/my/inventory/files?error=invalid_format')
            
            if result.get('error'):
                return request.redirect('/my/inventory/files?error=' + result['error'])
            
            return request.redirect(f'/my/inventory/files?imported={result.get("count", 0)}')
            
        except Exception as e:
            return request.redirect('/my/inventory/files?error=' + str(e))

    def _parse_file_xlsx(self, file_content):
        """Parse file folder data from XLSX file"""
        import io
        try:
            import openpyxl
        except ImportError:
            return {'error': 'Excel support not available'}
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            worksheet = workbook.active
            
            partner = request.env.user.partner_id
            File = request.env['records.file']
            Container = request.env['records.container']
            Department = request.env['records.department']
            
            created_count = 0
            
            # Skip header rows
            for row in worksheet.iter_rows(min_row=3, values_only=True):
                name = row[0]
                if not name:
                    continue
                
                container_name = row[1] if len(row) > 1 else ''
                description = row[2] if len(row) > 2 else ''
                department_name = row[3] if len(row) > 3 else ''
                
                # Find container by name
                container_id = False
                if container_name:
                    container = Container.sudo().search([
                        ('name', 'ilike', container_name),
                        ('partner_id', '=', partner.id)
                    ], limit=1)
                    if container:
                        container_id = container.id
                
                # Find department by name
                department_id = False
                if department_name:
                    dept = Department.sudo().search([
                        ('name', 'ilike', department_name),
                        ('company_id', '=', partner.commercial_partner_id.id)
                    ], limit=1)
                    if dept:
                        department_id = dept.id
                
                # Create file
                vals = {
                    'name': str(name),
                    'description': str(description) if description else '',
                    'container_id': container_id,
                    'created_via_portal': True,
                }
                
                File.sudo().create(vals)
                created_count += 1
            
            return {'count': created_count}
            
        except Exception as e:
            return {'error': str(e)}

    def _parse_file_csv(self, file_content):
        """Parse file folder data from CSV file"""
        import csv
        import io
        
        try:
            content = file_content.decode('utf-8')
            reader = csv.reader(io.StringIO(content))
            
            partner = request.env.user.partner_id
            File = request.env['records.file']
            Container = request.env['records.container']
            
            created_count = 0
            
            # Skip header rows
            next(reader, None)
            next(reader, None)
            
            for row in reader:
                if not row or not row[0]:
                    continue
                
                name = row[0]
                container_name = row[1] if len(row) > 1 else ''
                description = row[2] if len(row) > 2 else ''
                
                # Find container by name
                container_id = False
                if container_name:
                    container = Container.sudo().search([
                        ('name', 'ilike', container_name),
                        ('partner_id', '=', partner.id)
                    ], limit=1)
                    if container:
                        container_id = container.id
                
                # Create file
                vals = {
                    'name': name,
                    'description': description,
                    'container_id': container_id,
                    'created_via_portal': True,
                }
                
                File.sudo().create(vals)
                created_count += 1
            
            return {'count': created_count}
            
        except Exception as e:
            return {'error': str(e)}

    @http.route(['/my/users/import'], type='http', auth="user", website=True, methods=['POST'], csrf=True)
    def import_users(self, **post):
        """Import portal users from uploaded file (company admin only)"""
        # Security: Only company admins can import users
        if not request.env.user.has_group('records_management.group_portal_company_admin'):
            return request.redirect('/my/home')
        
        try:
            file_data = post.get('file')
            if not file_data:
                return request.redirect('/my/users?error=no_file')
            
            # Read file content
            file_content = file_data.read()
            
            # Determine file type and parse
            if file_data.filename.endswith('.xlsx'):
                result = self._parse_user_xlsx(file_content)
            elif file_data.filename.endswith('.csv'):
                result = self._parse_user_csv(file_content)
            else:
                return request.redirect('/my/users?error=invalid_format')
            
            if result.get('error'):
                return request.redirect('/my/users?error=' + result['error'])
            
            return request.redirect(f'/my/users?imported={result.get("count", 0)}')
            
        except Exception as e:
            return request.redirect('/my/users?error=' + str(e))

    def _parse_user_xlsx(self, file_content):
        """Parse user data from XLSX file"""
        import io
        try:
            import openpyxl
        except ImportError:
            return {'error': 'Excel support not available'}
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            worksheet = workbook.active
            
            partner = request.env.user.partner_id.commercial_partner_id
            User = request.env['res.users']
            Partner = request.env['res.partner']
            Department = request.env['records.department']
            portal_group = request.env.ref('base.group_portal')
            
            created_count = 0
            
            # Skip header rows
            for row in worksheet.iter_rows(min_row=3, values_only=True):
                name = row[0]
                email = row[1] if len(row) > 1 else ''
                department_name = row[2] if len(row) > 2 else ''
                phone = row[3] if len(row) > 3 else ''
                job_title = row[4] if len(row) > 4 else ''
                
                if not name or not email or not department_name:
                    continue
                
                # Check if user already exists
                existing = User.sudo().search([('login', '=', email)], limit=1)
                if existing:
                    continue
                
                # Find department by name
                dept = Department.sudo().search([
                    ('name', 'ilike', department_name),
                    ('company_id', '=', partner.id)
                ], limit=1)
                
                if not dept:
                    continue  # Skip if department not found
                
                # Create partner
                new_partner = Partner.sudo().create({
                    'name': str(name),
                    'email': str(email),
                    'parent_id': partner.id,
                    'type': 'contact',
                    'department_id': dept.id,
                    'phone': str(phone) if phone else '',
                    'function': str(job_title) if job_title else '',
                })
                
                # Create portal user
                new_user = User.sudo().create({
                    'name': str(name),
                    'login': str(email),
                    'email': str(email),
                    'partner_id': new_partner.id,
                    'groups_id': [(6, 0, [portal_group.id])],
                    'active': True,
                })
                
                # Send invitation email
                new_user.sudo().action_reset_password()
                
                created_count += 1
            
            return {'count': created_count}
            
        except Exception as e:
            return {'error': str(e)}

    def _parse_user_csv(self, file_content):
        """Parse user data from CSV file"""
        import csv
        import io
        
        try:
            content = file_content.decode('utf-8')
            reader = csv.reader(io.StringIO(content))
            
            partner = request.env.user.partner_id.commercial_partner_id
            User = request.env['res.users']
            Partner = request.env['res.partner']
            Department = request.env['records.department']
            portal_group = request.env.ref('base.group_portal')
            
            created_count = 0
            
            # Skip header rows
            next(reader, None)
            next(reader, None)
            
            for row in reader:
                if len(row) < 3 or not row[0] or not row[1] or not row[2]:
                    continue
                
                name = row[0]
                email = row[1]
                department_name = row[2]
                phone = row[3] if len(row) > 3 else ''
                job_title = row[4] if len(row) > 4 else ''
                
                # Check if user already exists
                existing = User.sudo().search([('login', '=', email)], limit=1)
                if existing:
                    continue
                
                # Find department by name
                dept = Department.sudo().search([
                    ('name', 'ilike', department_name),
                    ('company_id', '=', partner.id)
                ], limit=1)
                
                if not dept:
                    continue
                
                # Create partner
                new_partner = Partner.sudo().create({
                    'name': name,
                    'email': email,
                    'parent_id': partner.id,
                    'type': 'contact',
                    'department_id': dept.id,
                    'phone': phone,
                    'function': job_title,
                })
                
                # Create portal user
                new_user = User.sudo().create({
                    'name': name,
                    'login': email,
                    'email': email,
                    'partner_id': new_partner.id,
                    'groups_id': [(6, 0, [portal_group.id])],
                    'active': True,
                })
                
                # Send invitation email
                new_user.sudo().action_reset_password()
                
                created_count += 1
            
            return {'count': created_count}
            
        except Exception as e:
            return {'error': str(e)}
